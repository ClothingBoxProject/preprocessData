import re
import time
import hashlib
from pathlib import Path

import requests
from openpyxl import load_workbook

UNSPLASH_ACCESS_KEY = "eqBM1VfsIji-SoozJiq3bwN5gCr6AkRSxMN4BMCvx8E" # 환경변수로 넣는 걸 추천
API_URL = "https://api.unsplash.com/search/photos"

def safe_name(s: str, max_len: int = 80) -> str:
    s = str(s).strip()
    s = re.sub(r'[\\/:*?"<>|]+', "_", s)
    s = re.sub(r"\s+", " ", s)
    return s[:max_len]


def read_item_names_from_xlsx_range(
    xlsx_path: str,
    col_name: str = "item_name",
    sheet_name: str | None = None,
    start_excel_row: int = 2,  
):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    # 헤더는 1행이라고 가정
    header = [c.value for c in ws[1]]
    header_norm = [str(x).strip() if x is not None else "" for x in header]

    if col_name not in header_norm:
        raise RuntimeError(f"'{col_name}' 컬럼을 찾지 못했습니다. 현재 헤더: {header_norm}")

    idx = header_norm.index(col_name) + 1  # openpyxl은 1-based column

    items = []
    # ✅ start_excel_row ~ 마지막행까지
    for r in range(start_excel_row, ws.max_row + 1):
        v = ws.cell(row=r, column=idx).value
        if v is None:
            continue
        v = str(v).strip()
        if v:
            items.append((r, v))  # (엑셀행번호, item_name)

    return items


def search_unsplash_images(query: str, per_page: int = 30, page: int = 1) -> list[dict]:
    if not UNSPLASH_ACCESS_KEY:
        raise RuntimeError("환경변수 UNSPLASH_ACCESS_KEY 를 설정하세요.")

    params = {"query": query, "page": page, "per_page": per_page, "lang": "ko"}
    headers = {
        "Authorization": f"Client-ID {UNSPLASH_ACCESS_KEY}",
        "Accept-Version": "v1",
        "User-Agent": "Mozilla/5.0",
    }

    r = requests.get(API_URL, params=params, headers=headers, timeout=30)
    r.raise_for_status()
    data = r.json()

    results = []
    for item in data.get("results", []):
        results.append({
            "id": item.get("id"),
            "img_url": item["urls"].get("regular") or item["urls"].get("small") or item["urls"].get("full"),
        })
    return results


def download_image(url: str, out_path: Path, sleep: float = 0.2):
    out_path.parent.mkdir(parents=True, exist_ok=True)
    rr = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, stream=True, timeout=60)
    rr.raise_for_status()
    with open(out_path, "wb") as f:
        for chunk in rr.iter_content(chunk_size=1024 * 64):
            if chunk:
                f.write(chunk)
    time.sleep(sleep)


def crawl_from_xlsx_from_row50(
    xlsx_path: str,
    out_dir: str = "unsplash_photos",
    col_name: str = "item_name",
    sheet_name: str | None = None,
    start_excel_row: int = 50,  
    photos_per_item: int = 10,
    per_page: int = 30,
    max_pages: int = 5,
):
    items = read_item_names_from_xlsx_range(
        xlsx_path=xlsx_path,
        col_name=col_name,
        sheet_name=sheet_name,
        start_excel_row=start_excel_row,
    )
    print(f"[INFO] 처리 대상: {len(items)}개 (엑셀 {start_excel_row}행부터)")

    base = Path(out_dir)
    base.mkdir(parents=True, exist_ok=True)

    for i, (excel_row, item_name) in enumerate(items, 1):
        q = item_name
        folder = base / safe_name(q)

        saved = 0
        used_ids = set()

        print(f"\n[{i}/{len(items)}] excel_row={excel_row} query='{q}'")

        for page in range(1, max_pages + 1):
            results = search_unsplash_images(q, per_page=per_page, page=page)
            if not results:
                break

            for it in results:
                if saved >= photos_per_item:
                    break
                pid = it.get("id")
                url = it.get("img_url")
                if not pid or not url or pid in used_ids:
                    continue
                used_ids.add(pid)

                h = hashlib.sha1(url.encode("utf-8")).hexdigest()[:8]
                out_path = folder / f"{saved+1:02d}_{pid}_{h}.jpg"

                if not out_path.exists():
                    try:
                        download_image(url, out_path)
                        saved += 1
                        print(f"  [OK] {saved}/{photos_per_item} -> {out_path.name}")
                    except Exception as e:
                        print(f"  [FAIL] {url} : {e}")

            if saved >= photos_per_item:
                break

        if saved == 0:
            print("  [WARN] 저장된 이미지 없음")


if __name__ == "__main__":
    xlsx_path = "./filtered_postprocessed.xlsx"  
    out_dir = "./unsplash_photos"  

    crawl_from_xlsx_from_row50(
        xlsx_path=xlsx_path,
        out_dir=out_dir,
        col_name="item_name",
        sheet_name=None,
        start_excel_row=50,    
        photos_per_item=10,
        max_pages=5,
    )